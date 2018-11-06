import logging
import openpyxl
import sqlite3
import pandas as pd
import os
import re



def column_counter(xls_filename): #checks row 1 and counts how many cells have data, therefore how many columns in xls
    logging.debug('Counting columns in xls')
    wb = openpyxl.load_workbook(xls_filename)
    sheet = wb.active
    cols = 1 # start from 1st column
    while 1:
        cell = sheet.cell(row = 1, column = cols)
        v = cell.value
        if v != None: # if there is data in the cell
            cols += 1 # check the next column along
        else:    # if no data in the cell, then that's the last column, so break
            break
    c = int(cols)-1  # need to be minus one because it increments cols, then realises it's an empty cell
    logging.debug(f'# cols = {c}')
    return c


def row_counter(xls_filename): #checks column 1 and counts how many cells have data, therefore how many rows in xls
    logging.debug('Counting rows in xls')
    wb = openpyxl.load_workbook(xls_filename)
    sheet = wb.active
    rows = 1  # start from 1st column
    while 1:
        check_cell = sheet.cell(row = rows, column = 1)
        v = check_cell.value
        if v != None:  #if there is data in the cell
            rows += 1  # check the next column along
        else:    # if no data in the cell, then that's the last row, so break
             break
    r = int(rows)-1 # need to be minus one because it increments rows, then realises it's an empty cell
    logging.debug(f'#rows = {r}')
    return r


def excel_headings_grabber(xls_filename): # checks row 1 of xls and returns a dictionary showing col# & heading
    logging.debug('excel_headings_grabber - establishing headings/columns dict')
    wb = openpyxl.load_workbook(xls_filename)
    sheet = wb.active
    cols = 1
    dic = {}  # start from 1st column
    while 1:
        check_cell = sheet.cell(row = 1, column = cols)
        v = check_cell.value
        if v != None:  # if there is data in the cell
            dic.setdefault(cols, v)
            cols += 1 # check the next column along
        else:    # if no data in the cell, then that's the last column, so break
            break
    return dic


#### these functions are all copy/pasted, aim is to add Primary Key to existing database

def get_create_table_string(tablename, connection):
    sql = """
    select * from sqlite_master where name = "{}" and type = "table"
    """.format(tablename)
    result = connection.execute(sql)

    create_table_string = result.fetchmany()[0][4]
    return create_table_string


def add_pk_to_create_table_string(create_table_string, colname):
    regex = "(\n.+{}[^,]+)(,)".format(colname)
    return re.sub(regex, "\\1 PRIMARY KEY,",  create_table_string)


def add_pk_to_sqlite_table(tablename, index_column, connection):
    cts = get_create_table_string(tablename, connection)
    cts = add_pk_to_create_table_string(cts, index_column)
    template = """
    BEGIN TRANSACTION;
        ALTER TABLE {tablename} RENAME TO {tablename}_old_;
        
        {cts};
        
        INSERT INTO {tablename} SELECT * FROM {tablename}_old_;
        
        DROP TABLE {tablename}_old_;
        
    COMMIT TRANSACTION;
    """

    create_and_drop_sql = template.format(tablename = tablename, cts = cts)
    connection.executescript(create_and_drop_sql)

################



# use this part instead to import an excel file and use to create sqlite db with primary key
# unfortunately this just adds a new (numerical) primary key column rather than being able to set an existing col as PK

"""
desired_dir_1 = "H:\WorkingDir\member_data"
filename_1 = "All members.xlsx"
file_1 = "All members"

desired_dir_2 = "C:\Github local repos\excel_matcher\private"
filename_2 = "surveyresults.xlsx"
file_2 = "SurveyResults"

os.chdir(desired_dir_1)



xls_file = pd.ExcelFile(filename_1)
df = xls_file.parse(file_1)
con = sqlite3.connect(file_1 + ".db")
df.to_sql("df", con, if_exists="replace")
add_pk_to_sqlite_table("df", "index", con)
r = con.execute("select sql from sqlite_master where name = 'df' and type = 'table'")
print(r.fetchone()[0])

"""




# this code experiments with attaching SurveyResults DB to Main and reading from the databases
"""

dir = r"H:\WorkingDir\member_data"
os.chdir(dir)

All_members_db_file = "All members.db"

id_number = "70037A57-1226-43BB-8597-008B4AADA083"

conn = sqlite3.connect(All_members_db_file)
c = conn.cursor()

c.execute("ATTACH DATABASE 'SurveyResults.db' AS SurveyResults") # attaches SurveyResults db to the connection, which I later access via read_from_db2()


def read_from_db(member_id):
    table_name = "All_Members"
    id_column = "Id"
    age_column = "Age"
    gender_column = "Gender"
    c.execute("SELECT {rcn1}, {rcn2} FROM {tn} WHERE {lcn}='{idno}'".format(rcn1=age_column, rcn2=gender_column, tn=table_name, lcn=id_column, idno=member_id))
    return c.fetchone()


def read_from_db2(member_id):
    table_name = "SurveyResults"
    id_column = "Memberid"
    outcome_column = "Outcome"
    c.execute("SELECT {rcn1} FROM {tn} WHERE {lcn}='{idno}'".format(rcn1=outcome_column, tn=table_name, lcn=id_column, idno=member_id))
    print(c.fetchone())
    # return c.fetchone()


read_from_db2("d80488c3-08a7-4c7e-8777-a6d20060b56d")

# age, gender = read_from_db(id_number)
# print(age)
# print(gender)

"""




# test_member_IDs = ["277DA8F4-E360-46B2-8C29-A92100B685C9",
# "52EA3A2B-EFC4-42C5-B0A4-A201005B30EA",
# "D80488C3-08A7-4C7E-8777-A6D20060B56D",
# "895AC6E8-16DD-4FB9-833E-A3420089B3EA",
# "34D36CEE-3BFB-47AB-BF38-A753004B2B9F"

#
# test_member_IDs_5_to_10 = ["B2E84D86-40D9-47C2-85F0-A88500977BF3",
# "15047D34-D929-4A41-AB81-A951002C3171",
# "39E3B8DF-6CCA-4C06-812B-05A67B92222D",
# "87E19119-0245-448D-9568-6D75826CFA0F",
# "FA10AF87-2811-4B6B-A1F4-A30A00BC3D0E"]








# c.execute("SELECT * from {tn} WHERE {cn} = '277DA8F4-E360-46B2-8C29-A92100B685C9'".format(tn="All_Members", cn="Id"))
# data = c.fetchall()
# print(data)
